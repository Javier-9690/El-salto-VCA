"""
Comparador de Habitaciones - El Salto vs Hotelería
Flask app con PostgreSQL para persistir el mapa de habitaciones y los reportes.
"""

import os
import io
import uuid
import json
import base64
import traceback
import psycopg2
import psycopg2.extras
import pandas as pd
from flask import Flask, render_template, request, send_file, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32 MB


# ─────────────────────────────────────────────
#  Base de datos PostgreSQL
# ─────────────────────────────────────────────

def get_db():
    """Abre una conexión a PostgreSQL usando DATABASE_URL."""
    url = os.environ.get('DATABASE_URL', '')
    if not url:
        raise RuntimeError("Variable DATABASE_URL no configurada.")
    # Render usa postgres://, psycopg2 necesita postgresql://
    if url.startswith('postgres://'):
        url = url.replace('postgres://', 'postgresql://', 1)
    return psycopg2.connect(url, sslmode='require')


def init_db():
    """Crea las tablas si no existen."""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS configuracion (
                    clave       TEXT PRIMARY KEY,
                    valor       TEXT,
                    updated_at  TIMESTAMP DEFAULT NOW()
                );
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS comparaciones (
                    token       TEXT PRIMARY KEY,
                    excel_data  BYTEA NOT NULL,
                    created_at  TIMESTAMP DEFAULT NOW()
                );
            """)
            # Limpiar comparaciones con más de 7 días
            cur.execute("""
                DELETE FROM comparaciones
                WHERE created_at < NOW() - INTERVAL '7 days';
            """)
        conn.commit()


def guardar_mapa_db(df):
    """Guarda el DataFrame del mapa como JSON en la base de datos."""
    data = df.to_json(orient='records', force_ascii=False)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO configuracion (clave, valor, updated_at)
                VALUES ('mapa', %s, NOW())
                ON CONFLICT (clave) DO UPDATE
                    SET valor = EXCLUDED.valor,
                        updated_at = NOW();
            """, (data,))
        conn.commit()


def cargar_mapa_db():
    """Carga el mapa desde la base de datos. Retorna DataFrame o None."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT valor, updated_at FROM configuracion WHERE clave = 'mapa';"
                )
                row = cur.fetchone()
                if row and row[0]:
                    df = pd.DataFrame(json.loads(row[0]))
                    df = df.fillna('').astype(str)
                    return df, row[1]
    except Exception:
        pass
    return None, None


def guardar_excel_db(token, excel_bytes):
    """Guarda el Excel en la base de datos."""
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO comparaciones (token, excel_data, created_at)
                VALUES (%s, %s, NOW())
                ON CONFLICT (token) DO UPDATE
                    SET excel_data = EXCLUDED.excel_data,
                        created_at = NOW();
            """, (token, psycopg2.Binary(excel_bytes)))
        conn.commit()


def cargar_excel_db(token):
    """Carga el Excel desde la base de datos. Retorna bytes o None."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT excel_data FROM comparaciones WHERE token = %s;",
                    (token,)
                )
                row = cur.fetchone()
                if row:
                    return bytes(row[0])
    except Exception:
        pass
    return None


# Inicializar BD al arrancar
try:
    init_db()
except Exception as e:
    print(f"[WARN] No se pudo inicializar la BD: {e}")


# ─────────────────────────────────────────────
#  Utilidades
# ─────────────────────────────────────────────

def limpiar(val):
    if pd.isna(val):
        return ''
    return str(val).strip()


def norm_rut(val):
    if pd.isna(val):
        return ''
    return str(val).strip().upper().replace('.', '').replace(' ', '')


def quitar_tildes(texto):
    reemplazos = {
        'Á':'A','É':'E','Í':'I','Ó':'O','Ú':'U','Ü':'U','Ñ':'N',
        'á':'a','é':'e','í':'i','ó':'o','ú':'u','ü':'u','ñ':'n',
    }
    for k, v in reemplazos.items():
        texto = texto.replace(k, v)
    return texto


def normalizar_col(texto):
    return quitar_tildes(str(texto).strip().upper().replace('\n', ' ').replace('  ', ' '))


def buscar_col(df, nombres):
    mapa = {normalizar_col(c): c for c in df.columns}
    for nombre in nombres:
        if normalizar_col(nombre) in mapa:
            return mapa[normalizar_col(nombre)]
    return None


def leer_excel(data_bytes):
    """Lee .xls o .xlsx detectando la fila de encabezados automáticamente."""
    magic  = data_bytes[:4]
    engine = 'openpyxl' if magic[:2] == b'PK' else 'xlrd'

    df_raw = pd.read_excel(io.BytesIO(data_bytes), dtype=str,
                           engine=engine, header=None).fillna('')

    header_row = 0
    for i in range(min(5, len(df_raw))):
        celdas = [c for c in df_raw.iloc[i]
                  if str(c).strip() and len(str(c).strip()) <= 40]
        if len(celdas) >= 3:
            header_row = i
            break

    df = pd.read_excel(io.BytesIO(data_bytes), dtype=str,
                       engine=engine, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.fillna('')
    df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)]
    return df


# ─────────────────────────────────────────────
#  Procesamiento principal
# ─────────────────────────────────────────────

def procesar(mapa_bytes_o_df, salto_bytes, hotel_bytes):
    """
    Procesa los tres archivos.
    mapa_bytes_o_df puede ser bytes (archivo subido) o DataFrame (desde BD).
    """
    if isinstance(mapa_bytes_o_df, pd.DataFrame):
        df_map = mapa_bytes_o_df
    else:
        df_map = leer_excel(mapa_bytes_o_df)

    df_sal = leer_excel(salto_bytes)
    df_hot = leer_excel(hotel_bytes)

    for df in [df_map, df_sal, df_hot]:
        df.columns = [str(c).strip() for c in df.columns]
        df.fillna('', inplace=True)

    # ── Columnas Mapa ──────────────────────────────────────────────
    map_hab  = buscar_col(df_map, ['HABITACIÓN', 'HABITACION', 'HAB'])
    map_nm   = buscar_col(df_map, ['NM SALTO', 'NM_SALTO', 'NMSALTO'])
    map_camp = buscar_col(df_map, ['CAMPAMENTO'])
    map_mod  = buscar_col(df_map, ['MÓDULO', 'MODULO'])
    map_piso = buscar_col(df_map, ['PISO'])

    # ── Columnas El Salto ──────────────────────────────────────────
    sal_ext  = buscar_col(df_sal, ['ExtID', 'EXTID', 'EXT ID'])
    sal_door = buscar_col(df_sal, ['NameDoorList', 'NAMEDOORLIST', 'NAME DOOR LIST'])
    sal_name = buscar_col(df_sal, ['FullName', 'FULLNAME', 'FULL NAME'])

    # ── Columnas Hotelería ─────────────────────────────────────────
    hot_hab   = buscar_col(df_hot, ['HABITACIÓN', 'HABITACION', 'HAB',
                                     'HABITACION ', 'N° HAB', 'N°HAB', 'NRO HAB',
                                     'NUMERO HABITACION', 'NUMERO HABITACIÓN'])
    hot_rut   = buscar_col(df_hot, ['RUT', 'RUT TRABAJADOR', 'RUT_TRABAJADOR',
                                     'RUTTRABAJADOR', 'RUT PERSONA', 'DNI'])
    hot_nom   = buscar_col(df_hot, ['NOMBRE', 'NOMBRE COMPLETO', 'NOMBRES'])
    hot_emp   = buscar_col(df_hot, ['EMPRESA'])
    hot_mod   = buscar_col(df_hot, ['MÓDULO', 'MODULO'])
    hot_cont  = buscar_col(df_hot, ['N°CONTRATO', 'N CONTRATO', 'NCONTRATO',
                                     'NUMERO CONTRATO', 'N° CONTRATO'])
    hot_ger   = buscar_col(df_hot, ['GERENCIA'])
    hot_turno = buscar_col(df_hot, ['SISTEMA TURNO', 'SISTEMATURNO', 'TURNO',
                                     'SISTEMA\nTURNO', 'SISTEMA_TURNO'])

    faltantes = []
    if not map_hab:  faltantes.append("HABITACIÓN  →  Mapa de habitaciones")
    if not map_nm:   faltantes.append("NM SALTO    →  Mapa de habitaciones")
    if not sal_door: faltantes.append("NameDoorList →  Base de datos El Salto")
    if not hot_hab:  faltantes.append("HABITACIÓN  →  Base de datos Hotelería")
    if not hot_rut:  faltantes.append("RUT         →  Base de datos Hotelería")
    if faltantes:
        cols_hot = list(df_hot.columns)
        cols_sal = list(df_sal.columns)
        cols_map = list(df_map.columns)
        raise ValueError(
            "Columnas no encontradas:\n" + "\n".join(faltantes) +
            f"\n\n── Columnas detectadas en Hotelería ──\n{cols_hot}" +
            f"\n\n── Columnas detectadas en El Salto ──\n{cols_sal}" +
            f"\n\n── Columnas detectadas en Mapa ──\n{cols_map}"
        )

    # ── Mapeo bidireccional de habitaciones ───────────────────────
    h2n, n2h = {}, {}
    for _, fila in df_map.iterrows():
        h = limpiar(fila.get(map_hab, '')).upper()
        n = limpiar(fila.get(map_nm,  '')).upper()
        if h and n:
            h2n[h] = n
            n2h[n] = limpiar(fila.get(map_hab, ''))

    # ── Normalizar ────────────────────────────────────────────────
    df_hot['_RUT']   = df_hot[hot_rut].apply(norm_rut)
    df_sal['_RUT']   = df_sal[sal_ext].apply(norm_rut) if sal_ext else ''
    df_hot['_HAB']   = df_hot[hot_hab].apply(lambda x: limpiar(x).upper())
    df_sal['_DOOR']  = df_sal[sal_door].apply(lambda x: limpiar(x).upper())
    df_hot['_NM_EQ'] = df_hot['_HAB'].map(h2n)
    df_sal['_HAB_EQ']= df_sal['_DOOR'].map(n2h)

    hot_idx = {r['_RUT']: r for _, r in df_hot.iterrows() if r['_RUT']}
    sal_idx = {r['_RUT']: r for _, r in df_sal.iterrows() if r['_RUT']}

    comunes    = sorted(set(hot_idx) & set(sal_idx))
    solo_hot_k = sorted(set(hot_idx) - set(sal_idx))
    solo_sal_k = sorted(set(sal_idx) - set(hot_idx))

    discrepancias, coincidencias = [], []
    for rut in comunes:
        h = hot_idx[rut];  s = sal_idx[rut]
        nm_eq  = limpiar(h.get('_NM_EQ',  ''))
        door   = limpiar(s['_DOOR'])
        hab_eq = limpiar(s.get('_HAB_EQ', ''))
        coincide = bool(nm_eq) and nm_eq.upper() == door.upper()
        rec = {
            'RUT':               rut,
            'Nombre Hotelería':  limpiar(h.get(hot_nom, '')) if hot_nom else '',
            'Nombre El Salto':   limpiar(s.get(sal_name, '')) if sal_name else '',
            'HAB Hotelería':     limpiar(h.get(hot_hab, '')),
            'HAB El Salto':      limpiar(s.get(sal_door, '')),
            'Equiv Hotel→Salto': nm_eq,
            'Equiv Salto→Hotel': hab_eq,
            'Empresa':           limpiar(h.get(hot_emp,  '')) if hot_emp  else '',
            'Módulo':            limpiar(h.get(hot_mod,  '')) if hot_mod  else '',
            'Gerencia':          limpiar(h.get(hot_ger,  '')) if hot_ger  else '',
        }
        (coincidencias if coincide else discrepancias).append(rec)

    solo_hotel = [{
        'RUT': rut,
        'Nombre':     limpiar(hot_idx[rut].get(hot_nom,  '')) if hot_nom  else '',
        'HABITACIÓN': limpiar(hot_idx[rut].get(hot_hab,  '')),
        'Empresa':    limpiar(hot_idx[rut].get(hot_emp,  '')) if hot_emp  else '',
        'Módulo':     limpiar(hot_idx[rut].get(hot_mod,  '')) if hot_mod  else '',
        'N°Contrato': limpiar(hot_idx[rut].get(hot_cont, '')) if hot_cont else '',
        'Gerencia':   limpiar(hot_idx[rut].get(hot_ger,  '')) if hot_ger  else '',
        'Turno':      limpiar(hot_idx[rut].get(hot_turno,'')) if hot_turno else '',
    } for rut in solo_hot_k]

    solo_salto = [{
        'RUT/ExtID':       rut,
        'Nombre':          limpiar(sal_idx[rut].get(sal_name, '')) if sal_name else '',
        'HAB El Salto':    limpiar(sal_idx[rut].get(sal_door, '')),
        'HAB Equivalente': limpiar(sal_idx[rut].get('_HAB_EQ', '')),
    } for rut in solo_sal_k]

    hab_sin_mapa  = sorted({limpiar(r.get(hot_hab,''))
        for _, r in df_hot.iterrows()
        if not r.get('_NM_EQ') and limpiar(r.get(hot_hab,''))})
    door_sin_mapa = sorted({limpiar(r.get(sal_door,''))
        for _, r in df_sal.iterrows()
        if not r.get('_HAB_EQ') and limpiar(r.get(sal_door,''))})

    return {
        'discrepancias': discrepancias,
        'coincidencias': coincidencias,
        'solo_hotel':    solo_hotel,
        'solo_salto':    solo_salto,
        'hab_sin_mapa':  hab_sin_mapa,
        'door_sin_mapa': door_sin_mapa,
        'stats': {
            'total_hotel':    len(df_hot),
            'total_salto':    len(df_sal),
            'total_mapa':     len(df_map),
            'coincidencias':  len(coincidencias),
            'discrepancias':  len(discrepancias),
            'solo_hotel':     len(solo_hotel),
            'solo_salto':     len(solo_salto),
            'hab_sin_mapa':   len(hab_sin_mapa),
            'door_sin_mapa':  len(door_sin_mapa),
        },
    }


# ─────────────────────────────────────────────
#  Generación de Excel de resultados
# ─────────────────────────────────────────────

ROJO    = PatternFill("solid", fgColor="FFB3B3")
VERDE   = PatternFill("solid", fgColor="B3FFB3")
NARANJA = PatternFill("solid", fgColor="FFE5B3")
AZUL    = PatternFill("solid", fgColor="B3D9FF")
GRIS    = PatternFill("solid", fgColor="E0E0E0")
HEADER  = PatternFill("solid", fgColor="1F3864")
FHEADER = Font(bold=True, color="FFFFFF", size=11)
FCELL   = Font(size=10)
ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left',   vertical='center', wrap_text=True)


def _ajustar_cols(ws):
    for col in ws.columns:
        ancho = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ancho + 4, 45)
    ws.row_dimensions[1].height = 30


def _escribir_hoja(ws, datos, fill_fila=None):
    if not datos:
        ws.append(["Sin registros"])
        ws['A1'].fill = GRIS
        return
    headers = list(datos[0].keys())
    ws.append(headers)
    for celda in ws[1]:
        celda.fill = HEADER; celda.font = FHEADER; celda.alignment = ALIGN_C
    for i, fila in enumerate(datos, start=2):
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=i, column=j, value=fila.get(h, ''))
            if fill_fila:
                c.fill = fill_fila; c.font = FCELL; c.alignment = ALIGN_L
    _ajustar_cols(ws)


def generar_excel(results):
    wb = Workbook()
    ws1 = wb.active;         ws1.title = "Discrepancias RUT"
    _escribir_hoja(ws1, results['discrepancias'], ROJO)
    ws2 = wb.create_sheet("Solo en Hotelería")
    _escribir_hoja(ws2, results['solo_hotel'], NARANJA)
    ws3 = wb.create_sheet("Solo en El Salto")
    _escribir_hoja(ws3, results['solo_salto'], AZUL)
    ws4 = wb.create_sheet("Coincidencias")
    _escribir_hoja(ws4, results['coincidencias'], VERDE)

    for titulo, lista in [("Sin mapa (Hotelería)", results['hab_sin_mapa']),
                          ("Sin mapa (El Salto)",  results['door_sin_mapa'])]:
        ws = wb.create_sheet(titulo)
        ws.append([f"Habitaciones sin equivalente"])
        ws['A1'].fill = HEADER; ws['A1'].font = FHEADER
        for item in lista:
            ws.append([item])
        _ajustar_cols(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
#  Plantillas vacías descargables
# ─────────────────────────────────────────────

PLANTILLAS = {
    'mapa': {
        'nombre': 'plantilla_mapa_habitaciones.xlsx',
        'columnas': ['HABITACIÓN', 'CAMPAMENTO', 'MÓDULO', 'PISO', 'NM SALTO'],
        'ejemplo': [
            ['HAB-101', 'CAMPAMENTO A', 'MÓDULO 1', '1', 'SALTO-101'],
            ['HAB-102', 'CAMPAMENTO A', 'MÓDULO 1', '1', 'SALTO-102'],
        ],
    },
    'salto': {
        'nombre': 'plantilla_base_el_salto.xlsx',
        'columnas': ['FullName', 'ExtID', 'DoorQty', 'ZoneQty', 'NameDoorList'],
        'ejemplo': [
            ['Juan Pérez',    '12345678-9', '1', '1', 'SALTO-101'],
            ['María González','98765432-1', '1', '1', 'SALTO-205'],
        ],
    },
    'hotel': {
        'nombre': 'plantilla_base_hoteleria.xlsx',
        'columnas': ['HABITACIÓN','MÓDULO','RUT','NOMBRE','EMPRESA',
                     'N°CONTRATO','GERENCIA','SISTEMA TURNO'],
        'ejemplo': [
            ['HAB-101','MÓDULO 1','12345678-9','Juan Pérez',
             'Empresa A','CONT-001','GERENCIA 1','A'],
            ['HAB-205','MÓDULO 2','98765432-1','María González',
             'Empresa B','CONT-002','GERENCIA 2','B'],
        ],
    },
}


def generar_plantilla(tipo):
    info = PLANTILLAS[tipo]
    wb   = Workbook()
    ws   = wb.active
    ws.title = "Datos"
    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    EJ_FILL  = PatternFill("solid", fgColor="DCE6F1")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    EJ_FONT  = Font(italic=True, color="555555", size=10)

    for col, nombre in enumerate(info['columnas'], 1):
        c = ws.cell(1, col, nombre)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    for i, fila in enumerate(info['ejemplo'], 2):
        for col, val in enumerate(fila, 1):
            c = ws.cell(i, col, val)
            c.fill = EJ_FILL; c.font = EJ_FONT
            c.alignment = Alignment(horizontal='left', vertical='center')

    for col_idx, nombre in enumerate(info['columnas'], 1):
        ancho = max(len(nombre), max(
            len(str(f[col_idx-1])) for f in info['ejemplo'])) + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = min(ancho, 40)
    ws.freeze_panes = 'A2'

    wi = wb.create_sheet("INSTRUCCIONES")
    wi['A1'] = "No cambies los nombres de las columnas. Las filas de ejemplo (azul) pueden borrarse."
    wi['A1'].font = Font(bold=True, color="C00000", size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
#  Rutas Flask
# ─────────────────────────────────────────────

@app.route('/', methods=['GET'])
def index():
    df_mapa, mapa_fecha = cargar_mapa_db()
    mapa_cargado = df_mapa is not None
    return render_template('index.html',
                           mapa_cargado=mapa_cargado,
                           mapa_fecha=mapa_fecha)


@app.route('/guardar-mapa', methods=['POST'])
def guardar_mapa_ruta():
    """Sube y guarda el mapa de habitaciones en la BD."""
    if 'mapa' not in request.files or not request.files['mapa'].filename:
        return redirect(url_for('index'))
    try:
        df = leer_excel(request.files['mapa'].read())
        guardar_mapa_db(df)
        return render_template('index.html',
                               mapa_cargado=True,
                               mapa_fecha='justo ahora',
                               ok_mapa='Mapa de habitaciones guardado correctamente.')
    except Exception as e:
        df_mapa, mapa_fecha = cargar_mapa_db()
        return render_template('index.html',
                               mapa_cargado=df_mapa is not None,
                               mapa_fecha=mapa_fecha,
                               error=f'Error al guardar el mapa: {e}')


@app.route('/borrar-mapa', methods=['POST'])
def borrar_mapa_ruta():
    """Elimina el mapa de la BD."""
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM configuracion WHERE clave = 'mapa';")
            conn.commit()
    except Exception:
        pass
    return redirect(url_for('index'))


@app.route('/procesar', methods=['POST'])
def procesar_ruta():
    # Salto y Hotelería son siempre obligatorios
    for campo in ['salto', 'hotel']:
        if campo not in request.files or not request.files[campo].filename:
            return render_template('index.html',
                                   error='Debes subir los archivos de El Salto y Hotelería.',
                                   **_mapa_ctx())

    try:
        salto_b = request.files['salto'].read()
        hotel_b = request.files['hotel'].read()

        # Mapa: usar el subido ahora o el guardado en BD
        if 'mapa' in request.files and request.files['mapa'].filename:
            mapa_src = request.files['mapa'].read()
            # Guardar también en BD para la próxima vez
            guardar_mapa_db(leer_excel(mapa_src))
        else:
            df_mapa, _ = cargar_mapa_db()
            if df_mapa is None:
                return render_template('index.html',
                                       error='No hay un Mapa de habitaciones guardado. '
                                             'Súbelo en la sección superior o adjúntalo aquí.',
                                       **_mapa_ctx())
            mapa_src = df_mapa

        results = procesar(mapa_src, salto_b, hotel_b)
        excel_b  = generar_excel(results)

        token = str(uuid.uuid4())
        guardar_excel_db(token, excel_b)

        return render_template('results.html', results=results, token=token)

    except Exception as e:
        return render_template('index.html',
                               error=f'Error al procesar los archivos: {e}\n\n{traceback.format_exc()}',
                               **_mapa_ctx())


def _mapa_ctx():
    df_mapa, mapa_fecha = cargar_mapa_db()
    return {'mapa_cargado': df_mapa is not None, 'mapa_fecha': mapa_fecha}


@app.route('/descargar/<token>')
def descargar(token):
    excel_b = cargar_excel_db(token)
    if not excel_b:
        return render_template('index.html',
                               error='El reporte expiró (7 días). Procesa los archivos nuevamente.',
                               **_mapa_ctx())
    return send_file(
        io.BytesIO(excel_b),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='reporte_diferencias.xlsx'
    )


@app.route('/plantilla/<tipo>')
def descargar_plantilla(tipo):
    if tipo not in PLANTILLAS:
        return "Plantilla no encontrada.", 404
    return send_file(
        generar_plantilla(tipo),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=PLANTILLAS[tipo]['nombre']
    )


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
